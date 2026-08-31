#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取 results.json，生成 Excel 报表与 HTML 可视化报表。"""
import json, os, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
RES = json.load(open(os.path.join(HERE, "results.json"), encoding="utf-8"))
RAW = json.load(open(os.path.join(HERE, "raw_data.json"), encoding="utf-8"))
# ETF 数据为可选（仅用于叠加图）：抓取被限流时允许缺失，报表照常生成
_etf_path = os.path.join(HERE, "etf_data.json")
try:
    ETF = json.load(open(_etf_path, encoding="utf-8"))
except Exception as _e:
    print(f"! 未读到 etf_data.json（{_e}），本次报表跳过 ETF 叠加图")
    ETF = {"etfs": {}, "index_map": {}}
ETF_LINE_COLORS = ["#ffb74d", "#ba68c8", "#4dd0e1", "#aed581"]
GEN_DATE = datetime.date.today().strftime("%Y-%m-%d")
ALPHA = RES["meta"]["alpha"]
END_DATE = RES["meta"].get("end", datetime.date.today().strftime("%Y%m%d"))
GEN_DATE = RES["meta"].get("gen_date", GEN_DATE)
END_DISP = f"{END_DATE[:4]}-{END_DATE[4:6]}-{END_DATE[6:]}" if len(END_DATE) == 8 else END_DATE

INDEX_ORDER = ["上证指数", "创业板指", "科创50"]

# ---------------- 低买高卖策略回测 ----------------
FEE = 0.003  # 每轮往返费用率（ETF 管理费+双边滑点/佣金近似）

ETF_MAP = {
    "上证指数": [
        {"etf": "510050", "etf_name": "华夏上证50ETF",
         "note": "流动性好；跟踪上证50，与综指有大盘股风格偏离"},
        {"etf": "510300", "etf_name": "华泰柏瑞沪深300ETF",
         "note": "流动性最好；跟踪沪深300，比上证50更贴近综指但非同一指数"},
    ],
    "创业板指": [{"etf": "159915", "etf_name": "易方达创业板ETF",
                  "note": "规模最大、流动性最好，直接跟踪创业板指"}],
    "科创50":   [{"etf": "588000", "etf_name": "华夏科创50ETF",
                  "note": "规模最大、流动性最好，直接跟踪科创50"}],
}

def compute_strategy():
    result = {"etfs": ETF_MAP, "indices": {}}
    for name in INDEX_ORDER:
        years = sorted(RES["indices"][name]["years"])
        trades = []
        for i, y in enumerate(years):
            vy = RES["indices"][name]["years"][y]
            # 跳过仅基点记录的年份（如 科创50 2019 只有基点 1000，非真实交易）
            if vy.get("trading_days", 999) < 30:
                continue
            if i == len(years) - 1:
                trades.append({
                    "buy_year": y, "sell_year": "持仓", "open": True,
                    "buy_date": vy["annual_low_date"], "buy_price_ext": vy["annual_low"],
                    "buy_band_mid": round((vy["low_band_low"] + vy["low_band_high"]) / 2, 2),
                    "sell_date": "—", "sell_price_ext": None, "sell_band_mid": None,
                    "r_ext": None, "r_mid": None,
                })
                continue
            vn = RES["indices"][name]["years"][years[i + 1]]
            buy_ext = vy["annual_low"]; sell_ext = vn["annual_high"]
            buy_mid = (vy["low_band_low"] + vy["low_band_high"]) / 2
            sell_mid = (vn["high_band_low"] + vn["high_band_high"]) / 2
            r_ext = (sell_ext - buy_ext) / buy_ext * 100
            r_mid = (sell_mid - buy_mid) / buy_mid * 100
            trades.append({
                "buy_year": y, "sell_year": years[i + 1], "open": False,
                "buy_date": vy["annual_low_date"], "buy_price_ext": buy_ext,
                "buy_band_mid": round(buy_mid, 2),
                "sell_date": vn["annual_high_date"], "sell_price_ext": sell_ext,
                "sell_band_mid": round(sell_mid, 2),
                "r_ext": round(r_ext, 2), "r_mid": round(r_mid, 2),
            })
        closed = [t for t in trades if not t["open"]]
        n = len(closed)
        avg_ext = sum(t["r_ext"] for t in closed) / n
        avg_mid = sum(t["r_mid"] for t in closed) / n
        avg_ext_net = sum(t["r_ext"] - FEE * 100 for t in closed) / n
        avg_mid_net = sum(t["r_mid"] - FEE * 100 for t in closed) / n
        result["indices"][name] = {
            "trades": trades, "n_closed": n,
            "avg_ext": round(avg_ext, 2), "avg_mid": round(avg_mid, 2),
            "avg_ext_net": round(avg_ext_net, 2), "avg_mid_net": round(avg_mid_net, 2),
        }
    ce = sum(result["indices"][n]["avg_ext"] for n in INDEX_ORDER) / 3
    cm = sum(result["indices"][n]["avg_mid"] for n in INDEX_ORDER) / 3
    cen = sum(result["indices"][n]["avg_ext_net"] for n in INDEX_ORDER) / 3
    cmn = sum(result["indices"][n]["avg_mid_net"] for n in INDEX_ORDER) / 3
    result["comprehensive"] = {
        "avg_ext": round(ce, 2), "avg_mid": round(cm, 2),
        "avg_ext_net": round(cen, 2), "avg_mid_net": round(cmn, 2),
    }
    return result

STRAT = compute_strategy()

def band_pcts(v):
    """计算价格带浮动百分比。"""
    low_pct = (v["low_band_high"] - v["low_band_low"]) / v["low_band_low"] * 100
    high_pct = (v["high_band_high"] - v["high_band_low"]) / v["high_band_low"] * 100
    mid_low = (v["low_band_low"] + v["low_band_high"]) / 2
    mid_high = (v["high_band_low"] + v["high_band_high"]) / 2
    band_to_band = (mid_high - mid_low) / mid_low * 100
    return round(low_pct, 2), round(high_pct, 2), round(band_to_band, 2)

# ---------------- Excel ----------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Side(style="thin", color="C9D6E5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, color="1F3A5F", size=14)
idx_font = Font(bold=True, color="0B6E4F" if False else "1F3A5F", size=12)

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

# 说明 sheet
ws0 = wb.active
ws0.title = "说明"
notes = [
    ("A股三大指数 逐年低点/高点区间报表", title_font),
    (f"生成日期：{GEN_DATE}", None),
    ("", None),
    ("数据来源：东方财富行情接口（日K线，前复权）", None),
    ("覆盖区间：2016-01-01 至 " + END_DISP + "（动态截止到最新交易日）", None),
    ("指数：上证指数(000001)、创业板指(399006)、科创50(000688)", None),
    ("", None),
    ("区间定义（与确认方案一致）：", Font(bold=True)),
    ("1) 年度最低点 = 全年每日最低价的最小值；年度最高点 = 全年每日最高价的最大值。", None),
    ("2) 低点区间 = 以「年最低点所在交易日」为锚，向两侧连续扩展，直到某日收盘价", None),
    (f"   首次突破 年最低×(1+{ALPHA:.0%}) 为止；取这段连续时段的起止日期与其间实际价格带。", None),
    ("3) 高点区间 = 对称处理，阈值 = 年最高×(1-5%)，直到收盘价跌破为止。", None),
    ("4) 仅取主区间（围绕年度极值的一段），若出现双底/双顶仅保留主区间。", None),
    ("", None),
    ("特别说明：", Font(bold=True)),
    ("· 科创50 指数于 2020-07-22 正式发布，2019 年仅有基点(1000)记录，故 2019 行", None),
    ("  仅显示基点，不代表真实交易区间；2016-2018 该指数尚未存在，无数据。", None),
    ("· 部分年份的高/低点为主板单日尖峰（如 2024-10-08），故区间仅 1 个交易日。", None),
    ("", None),
    ("免责声明：本报表仅基于公开行情数据的量化整理，不构成任何投资建议。", Font(bold=True, color="B00020")),
]
for i, (txt, fnt) in enumerate(notes, start=1):
    c = ws0.cell(row=i, column=1, value=txt)
    if fnt:
        c.font = fnt
ws0.column_dimensions["A"].width = 95

COLS = ["年份", "年度最低点", "最低点日期", "低点区间起", "低点区间止",
        "低点价格带(低~高)", "低点带浮动%", "年度最高点", "最高点日期", "高点区间起", "高点区间止",
        "高点价格带(低~高)", "高点带浮动%", "低→高上涨%区间", "交易日数"]

key_fill = PatternFill("solid", fgColor="2E4A22")
key_font = Font(bold=True, color="FFD54F", size=11)

for name in INDEX_ORDER:
    info = RES["indices"][name]
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=f"{name}（{info['code']}）逐年低点/高点区间").font = title_font
    ws.append([])
    ws.append(COLS)
    style_header(ws, 3, len(COLS))
    for y in sorted(info["years"]):
        v = info["years"][y]
        lp, hp, bp = band_pcts(v)
        ws.append([
            y, v["annual_low"], v["annual_low_date"], v["low_window_start"], v["low_window_end"],
            f"{v['low_band_low']}~{v['low_band_high']}", lp,
            v["annual_high"], v["annual_high_date"], v["high_window_start"], v["high_window_end"],
            f"{v['high_band_low']}~{v['high_band_high']}", hp, bp, v["trading_days"],
        ])
    # 样式
    for r in range(4, 4 + len(info["years"])):
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        # 高亮"低→高上涨%区间"列
        kc = ws.cell(row=r, column=14)
        kc.fill = key_fill
        kc.font = key_font
    widths = [8, 11, 12, 12, 12, 17, 11, 11, 12, 12, 12, 17, 11, 14, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

# 汇总 sheet
ws_all = wb.create_sheet("汇总对比")
ws_all.cell(row=1, column=1, value="三指数逐年低/高点对比（单位：点）").font = title_font
hdr = ["年份"]
for name in INDEX_ORDER:
    hdr += [f"{name}年最低", f"{name}年最高"]
ws_all.append([])
ws_all.append(hdr)
style_header(ws_all, 3, len(hdr))
years_all = sorted({y for name in INDEX_ORDER for y in RES["indices"][name]["years"]})
for y in years_all:
    row = [y]
    for name in INDEX_ORDER:
        v = RES["indices"][name]["years"].get(y)
        row += [v["annual_low"] if v else "-", v["annual_high"] if v else "-"]
    ws_all.append(row)
for r in range(4, 4 + len(years_all)):
    for c in range(1, len(hdr) + 1):
        ws_all.cell(row=r, column=c).border = border
        ws_all.cell(row=r, column=c).alignment = Alignment(horizontal="center")
for i in range(1, len(hdr) + 1):
    ws_all.column_dimensions[get_column_letter(i)].width = 13 if i == 1 else 14
ws_all.freeze_panes = "B4"

# 策略回测 sheet
def add_strategy_sheet():
    ws = wb.create_sheet("低买高卖回测")
    ws.cell(row=1, column=1, value="低买高卖策略回测（实际操盘参考）").font = title_font
    ws.cell(row=2, column=1, value=f"生成日期 {GEN_DATE} ｜ 规则：买入年低点区间买入→次年高点区间卖出(跨年持有)；极值=理论上限，区间中点=实操近似；每轮往返扣费 {FEE*100:.1f}%").font = Font(size=10, color="6F8BAB")

    r = 4
    ws.cell(row=r, column=1, value="一、代表 ETF 对照").font = Font(bold=True, size=12, color="1F3A5F"); r += 1
    ws.append([]) if False else None
    hdr = ["指数", "ETF代码", "ETF名称", "备注"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    r += 1
    for name in INDEX_ORDER:
        etfs = STRAT["etfs"][name]
        for j, e in enumerate(etfs):
            vals = [name if j == 0 else "", e["etf"], e["etf_name"], e["note"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v); cell.border = border
                cell.alignment = Alignment(horizontal="left" if c in (1, 4) else "center", vertical="center", wrap_text=True)
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="二、综合年盈利（等权平均，单位 %）").font = Font(bold=True, size=12, color="1F3A5F"); r += 1
    comp = STRAT["comprehensive"]
    comp_rows = [
        ("平均收益率·极值(理论上限)", comp["avg_ext"]),
        ("平均收益率·区间中点(实操近似)", comp["avg_mid"]),
        ("平均净收益·极值(扣费后)", comp["avg_ext_net"]),
        ("平均净收益·区间中点(扣费后)", comp["avg_mid_net"]),
    ]
    for c, h in enumerate(["指标", "数值(%)"], 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    r += 1
    for k, v in comp_rows:
        ws.cell(row=r, column=1, value=k).border = border
        cell = ws.cell(row=r, column=2, value=v); cell.border = border; cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, color="0B6E4F" if v >= 0 else "B00020")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="三、逐年交易明细（买年低点→次年高点）").font = Font(bold=True, size=12, color="1F3A5F"); r += 1
    tr_hdr = ["指数", "买入年", "买入日", "买入点(极值)", "买入点(中点)", "卖出年", "卖出日",
              "卖出点(极值)", "卖出点(中点)", "收益率%(极值)", "收益率%(中点)", "净收益%(中点,扣费)"]
    for c, h in enumerate(tr_hdr, 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
    r += 1
    for name in INDEX_ORDER:
        idx = STRAT["indices"][name]
        for t in idx["trades"]:
            row = [
                name, t["buy_year"], t["buy_date"], t["buy_price_ext"], t["buy_band_mid"],
                t["sell_year"], t["sell_date"],
                t["sell_price_ext"] if t["sell_price_ext"] is not None else "—",
                t["sell_band_mid"] if t["sell_band_mid"] is not None else "—",
                t["r_ext"] if t["r_ext"] is not None else "持仓",
                t["r_mid"] if t["r_mid"] is not None else "持仓",
                (round(t["r_mid"] - FEE * 100, 2)) if t["r_mid"] is not None else "—",
            ]
            for c, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=v); cell.border = border
                cell.alignment = Alignment(horizontal="center")
            r += 1
        # 该指数均值行
        row = [name + " 平均", "", "", "", "", "", "", "", "",
               idx["avg_ext"], idx["avg_mid"], idx["avg_mid_net"]]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="E8F0F7"); cell.font = Font(bold=True)
        r += 2

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["D"].width = 22  # 备注列

add_strategy_sheet()

# 当前位置 sheet
def add_live_sheet():
    ws = wb.create_sheet("当前位置")
    ws.cell(row=1, column=1, value=f"当前位置实时读数（截至 {RES['meta']['gen_date']}）").font = title_font
    ws.cell(row=2, column=1, value="以年初至今最低/最高收盘为锚，沿用 ±5% 容忍带；距低点带≤5%为低点带内，距高点带≥-5%为高点带内").font = Font(size=10, color="6F8BAB")
    hdr = ["指数", "状态", "今日收盘", "今日日期", "年内最低收盘", "最低日", "年内最高收盘", "最高日",
           "距低点带%", "距高点带%", "低点带阈值(≤5%)", "高点带阈值(≥-5%)"]
    r = 4
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
    r += 1
    for name in INDEX_ORDER:
        lv = RES["live"].get(name, {})
        if not lv:
            continue
        zone = lv["zone"]
        zfill = PatternFill("solid", fgColor="1EC98B" if zone == "低点带内" else ("FF5B5B" if zone == "高点带内" else "FFD54F"))
        vals = [name, zone, lv["last_close"], lv["last_date"], lv["ytd_low_close"], lv["ytd_low_date"],
                lv["ytd_high_close"], lv["ytd_high_date"], lv["dist_low_pct"], lv["dist_high_pct"],
                f"{lv['low_anchor']} (≤{ALPHA*100:.0f}%)", f"{lv['high_anchor']} (≥-{ALPHA*100:.0f}%)"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.border = border
            cell.alignment = Alignment(horizontal="center")
        zc = ws.cell(row=r, column=2)
        zc.fill = zfill; zc.font = Font(bold=True, color="0a1929")
        r += 1
    widths = [12, 12, 11, 12, 13, 12, 13, 12, 12, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

add_live_sheet()

xlsx_path = os.path.join(HERE, "index_low_high_report.xlsx")
wb.save(xlsx_path)
print("saved xlsx ->", xlsx_path)

# ---------------- HTML ----------------
LOW_COLOR = "#1EC98B"   # 低点区间（绿）
HIGH_COLOR = "#FF5B5B"  # 高点区间（红）

def svg_chart(name):
    info = RES["indices"][name]
    years = sorted(info["years"])
    raw = RAW[name]["series"]
    dates = [datetime.datetime.strptime(r["date"], "%Y-%m-%d").date() for r in raw]
    closes = [r["close"] for r in raw]
    lows = [r["low"] for r in raw]
    highs = [r["high"] for r in raw]
    dmin = min(dates).replace(month=1, day=1)
    dmax = max(dates).replace(month=12, day=31)
    pmin, pmax = min(lows), max(highs)
    pad = (pmax - pmin) * 0.05
    pLo, pHi = pmin - pad, pmax + pad

    DAY_W = 1.2  # px per calendar day => wide, scrollable chart
    left, right, top, bottom = 64, 40, 34, 70
    total_days = (dmax - dmin).days + 1
    plotW = total_days * DAY_W
    W = int(left + plotW + right)
    H = 420
    plotH = H - top - bottom

    def X(d):
        return left + (d.toordinal() - dmin.toordinal()) / total_days * plotW

    def Y(p):
        return top + plotH - (p - pLo) / (pHi - pLo) * plotH

    pts = " ".join(f"{X(d):.1f},{Y(c):.1f}" for d, c in zip(dates, closes))

    svg = [f'<svg viewBox="0 0 {W} {H}" width="{W}px" height="{H}px" style="min-width:{W}px" font-family="-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif">']

    # year separator lines + year labels (at top)
    for yr in range(dmin.year, dmax.year + 1):
        yd = datetime.date(yr, 1, 1)
        if yd < dmin or yd > dmax:
            continue
        x = X(yd)
        svg.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top+plotH}" stroke="#1c3a5e" stroke-width="1.2"/>')
        svg.append(f'<text x="{x+6:.1f}" y="{top-10}" fill="#b0c7e3" font-size="12" text-anchor="start" font-weight="600">{yr}年</text>')

    # monthly ticks + month numbers (1-12)
    cur = datetime.date(dmin.year, dmin.month, 1)
    while cur <= dmax:
        if cur >= dmin:
            x = X(cur)
            mid = x + (X(datetime.date(cur.year + (cur.month // 12), (cur.month % 12) + 1, 1)) - x) / 2
            svg.append(f'<line x1="{x:.1f}" y1="{top+plotH}" x2="{x:.1f}" y2="{top+plotH+6}" stroke="#2a4364" stroke-width="1"/>')
            svg.append(f'<text x="{mid:.1f}" y="{top+plotH+18}" fill="#7d97b6" font-size="9" text-anchor="middle">{cur.month}</text>')
        # advance 1 month
        y, m = cur.year, cur.month + 1
        if m > 12:
            y += 1
            m = 1
        cur = datetime.date(y, m, 1)

    # background bands for intervals
    for y in years:
        v = info["years"][y]
        ds = datetime.datetime.strptime(v["low_window_start"], "%Y-%m-%d").date()
        de = datetime.datetime.strptime(v["low_window_end"], "%Y-%m-%d").date()
        x0, x1 = X(ds), X(de)
        svg.append(f'<rect x="{x0:.1f}" y="{top}" width="{max(2, x1-x0):.1f}" height="{plotH}" fill="{LOW_COLOR}" opacity="0.14"/>')
        ds = datetime.datetime.strptime(v["high_window_start"], "%Y-%m-%d").date()
        de = datetime.datetime.strptime(v["high_window_end"], "%Y-%m-%d").date()
        x0, x1 = X(ds), X(de)
        svg.append(f'<rect x="{x0:.1f}" y="{top}" width="{max(2, x1-x0):.1f}" height="{plotH}" fill="{HIGH_COLOR}" opacity="0.14"/>')

    svg.append(f'<text x="{left}" y="{top+plotH+44}" fill="#8fa8c4" font-size="10" text-anchor="start">日期（下方数字为月份 1–12，可左右滑动）</text>')

    # y grid & labels
    for k in range(6):
        p = pLo + (pHi - pLo) * k / 5
        y = Y(p)
        svg.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left+plotW}" y2="{y:.1f}" stroke="#21364f" stroke-width="1"/>')
        svg.append(f'<text x="{left-8}" y="{y+3:.1f}" fill="#8fa8c4" font-size="11" text-anchor="end">{p:.0f}</text>')
    svg.append(f'<text x="{left-50}" y="{top-8}" fill="#8fa8c4" font-size="10" text-anchor="start">价格(点)</text>')

    # border
    svg.append(f'<rect x="{left}" y="{top}" width="{plotW}" height="{plotH}" fill="none" stroke="#1c3a5e" stroke-width="1.5"/>')

    # ---- ETF overlay (dual right axis) ----
    codes = ETF.get("index_map", {}).get(name, [])
    etf_series = []
    for code in codes:
        s = ETF.get("etfs", {}).get(code, {}).get("series", [])
        if not s:
            continue
        bd = {datetime.datetime.strptime(r["date"], "%Y-%m-%d").date(): r["close"] for r in s}
        etf_series.append((code, ETF["etfs"][code]["name"], bd))
    if etf_series:
        common = [d for d in dates if all(d in bd for _, _, bd in etf_series)]
        if common:
            allv = [bd[d] for _, _, bd in etf_series for d in common]
            emin, emax = min(allv), max(allv)
            epad = (emax - emin) * 0.12 or 0.1
            eLo, eHi = emin - epad, emax + epad
            def YE(v):
                return top + plotH - (v - eLo) / (eHi - eLo) * plotH
            for k in range(6):
                v = eLo + (eHi - eLo) * k / 5
                y = YE(v)
                svg.append(f'<line x1="{left+plotW:.1f}" y1="{y:.1f}" x2="{left+plotW+right-8:.1f}" y2="{y:.1f}" stroke="#21364f" stroke-width="1"/>')
                svg.append(f'<text x="{left+plotW+4:.1f}" y="{y+3:.1f}" fill="#8fa8c4" font-size="9" text-anchor="start">{v:.2f}</text>')
            svg.append(f'<text x="{W-2:.1f}" y="{top-8}" fill="#8fa8c4" font-size="10" text-anchor="end">ETF价格(元)</text>')
            for i, (code, nm, bd) in enumerate(etf_series):
                col = ETF_LINE_COLORS[i % len(ETF_LINE_COLORS)]
                pts_e = " ".join(f"{X(d):.1f},{YE(bd[d]):.1f}" for d in common)
                svg.append(f'<polyline points="{pts_e}" fill="none" stroke="{col}" stroke-width="1.3" stroke-dasharray="5 3" stroke-linejoin="round" opacity="0.95"/>')

    # trend line (index)
    svg.append(f'<polyline points="{pts}" fill="none" stroke="#89c4f4" stroke-width="1.4" stroke-linejoin="round"/>')

    # annual low / high dots (click -> popup detail; 不绑 hover，避免光标掠过时自动跳转表格)
    safe_name = name.replace(" ", "")
    for y in years:
        v = info["years"][y]
        dd = datetime.datetime.strptime(v["annual_low_date"], "%Y-%m-%d").date()
        x, ypos = X(dd), Y(v["annual_low"])
        # 小圆点半径仅 4px，太小不好点：外层再套一个透明可点击圆（与买/卖标记同一做法）
        svg.append(f'<circle cx="{x:.1f}" cy="{ypos:.1f}" r="4" fill="{LOW_COLOR}" stroke="#fff" stroke-width="1"/>')
        svg.append(f'<circle cx="{x:.1f}" cy="{ypos:.1f}" r="9" fill="transparent" stroke="none" pointer-events="all" style="cursor:pointer" onclick="showPopup(\'{safe_name}\', \'{y}\', \'low\')"><title>{v["annual_low_date"]} 年最低 {v["annual_low"]}（点击查看详情）</title></circle>')
        dd = datetime.datetime.strptime(v["annual_high_date"], "%Y-%m-%d").date()
        x, ypos = X(dd), Y(v["annual_high"])
        svg.append(f'<circle cx="{x:.1f}" cy="{ypos:.1f}" r="4" fill="{HIGH_COLOR}" stroke="#fff" stroke-width="1"/>')
        svg.append(f'<circle cx="{x:.1f}" cy="{ypos:.1f}" r="9" fill="transparent" stroke="none" pointer-events="all" style="cursor:pointer" onclick="showPopup(\'{safe_name}\', \'{y}\', \'high\')"><title>{v["annual_high_date"]} 年最高 {v["annual_high"]}（点击查看详情）</title></circle>')

    # ---- buy / sell action markers (at band midpoints) ----
    for y in years:
        v = info["years"][y]
        # low band midpoint -> 买 (green up-triangle)
        lm = (v["low_band_low"] + v["low_band_high"]) / 2
        ds = datetime.datetime.strptime(v["low_window_start"], "%Y-%m-%d").date()
        de = datetime.datetime.strptime(v["low_window_end"], "%Y-%m-%d").date()
        md = ds + (de - ds) // 2
        xb, yb = X(md), Y(lm)
        # 用透明圆形覆盖层把买/卖标记做成可点击区域，同时拦截底层年度极值点的 onmouseenter，
        # 避免光标移到标记上时自动跳转到下方表格；仅 click 触发弹窗。
        svg.append(f'<g class="marker-buy">')
        svg.append(f'<polygon points="{xb:.1f},{yb-6:.1f} {xb-5:.1f},{yb+4:.1f} {xb+5:.1f},{yb+4:.1f}" fill="{LOW_COLOR}" stroke="#fff" stroke-width="0.8"/>')
        svg.append(f'<text x="{xb:.1f}" y="{yb-10:.1f}" fill="{LOW_COLOR}" font-size="10" font-weight="700" text-anchor="middle">买</text>')
        svg.append(f'<circle cx="{xb:.1f}" cy="{yb:.1f}" r="10" fill="transparent" stroke="none" pointer-events="all" style="cursor:pointer" onclick="showPopup(\'{safe_name}\', \'{y}\', \'low\')"><title>{y}年低点区间买入信号（点击查看详情）</title></circle>')
        svg.append(f'</g>')
        # high band midpoint -> 卖 (red down-triangle)
        hm = (v["high_band_low"] + v["high_band_high"]) / 2
        ds = datetime.datetime.strptime(v["high_window_start"], "%Y-%m-%d").date()
        de = datetime.datetime.strptime(v["high_window_end"], "%Y-%m-%d").date()
        md = ds + (de - ds) // 2
        xs, ys = X(md), Y(hm)
        svg.append(f'<g class="marker-sell">')
        svg.append(f'<polygon points="{xs:.1f},{ys+6:.1f} {xs-5:.1f},{ys-4:.1f} {xs+5:.1f},{ys-4:.1f}" fill="{HIGH_COLOR}" stroke="#fff" stroke-width="0.8"/>')
        svg.append(f'<text x="{xs:.1f}" y="{ys+18:.1f}" fill="{HIGH_COLOR}" font-size="10" font-weight="700" text-anchor="middle">卖</text>')
        svg.append(f'<circle cx="{xs:.1f}" cy="{ys:.1f}" r="10" fill="transparent" stroke="none" pointer-events="all" style="cursor:pointer" onclick="showPopup(\'{safe_name}\', \'{y}\', \'high\')"><title>{y}年高点区间卖出信号（点击查看详情）</title></circle>')
        svg.append(f'</g>')

    # ---- 最新交易日（今日）标记 ----
    if dates:
        last_d, last_c = dates[-1], closes[-1]
        xl, yl = X(last_d), Y(last_c)
        svg.append(f'<line x1="{xl:.1f}" y1="{top}" x2="{xl:.1f}" y2="{top+plotH}" stroke="#FFD54F" stroke-width="1.2" stroke-dasharray="4 3" opacity="0.85"/>')
        svg.append(f'<circle cx="{xl:.1f}" cy="{yl:.1f}" r="3.5" fill="#FFD54F" stroke="#0c2238" stroke-width="1"/>')
        lbl = "今日" if last_d.year == dmax.year else f"{last_d.year}末"
        svg.append(f'<text x="{xl:.1f}" y="{top-10:.1f}" fill="#FFD54F" font-size="10" font-weight="700" text-anchor="middle">{lbl}</text>')

    svg.append('</svg>')
    return "".join(svg)

def html_table(name):
    info = RES["indices"][name]
    safe_name = name.replace(" ", "")
    rows = []
    for y in sorted(info["years"]):
        v = info["years"][y]
        lp, hp, bp = band_pcts(v)
        rows.append(f"""<tr id="row-{safe_name}-{y}">
<td>{y}</td>
<td class="num">{v['annual_low']}</td><td>{v['annual_low_date']}</td>
<td>{v['low_window_start']}</td><td>{v['low_window_end']}</td>
<td class="band">{v['low_band_low']} ~ {v['low_band_high']}</td>
<td class="pct">{lp}%</td>
<td class="num hi">{v['annual_high']}</td><td>{v['annual_high_date']}</td>
<td>{v['high_window_start']}</td><td>{v['high_window_end']}</td>
<td class="band">{v['high_band_low']} ~ {v['high_band_high']}</td>
<td class="pct">{hp}%</td>
<td class="key">{bp}%</td>
</tr>""")
    # ETF legend chips for the chart overlay
    codes = ETF.get("index_map", {}).get(name, [])
    etf_legend = ""
    if codes:
        parts = []
        for i, code in enumerate(codes):
            if code not in ETF.get("etfs", {}):
                continue
            col = ETF_LINE_COLORS[i % len(ETF_LINE_COLORS)]
            nm = ETF["etfs"][code]["name"]
            parts.append(f'<span class="line" style="display:inline-block;width:16px;height:0;border-top:2px dashed {col};vertical-align:middle;margin:0 3px"></span>{nm}({code})')
        etf_legend = "&nbsp;&nbsp;" + "&nbsp;&nbsp;".join(parts)
    caveat = ""
    if name == "科创50":
        caveat = '<p class="caveat">⚠ 科创50 于 2020-07-22 正式发布，2019 年仅基点(1000)记录；2016–2018 尚未存在，无数据。</p>'
    return f"""<div class="card" id="card-{safe_name}">
<h2>{name} <span class="code">({info['code']})</span></h2>
<div class="legend"><span class="dot low"></span>低点区间&nbsp;&nbsp;<span class="dot high"></span>高点区间&nbsp;&nbsp;<span class="line" style="display:inline-block;width:16px;height:2px;background:#89c4f4;vertical-align:middle;margin:0 3px"></span>指数收盘价&nbsp;&nbsp;{etf_legend}&nbsp;&nbsp;<span style="color:{LOW_COLOR};font-weight:700">▲买</span>&nbsp;(低带中点)&nbsp;&nbsp;<span style="color:{HIGH_COLOR};font-weight:700">▼卖</span>&nbsp;(高带中点)&nbsp;&nbsp;<span class="anno">● = 年度极值点（点击查看详情）；▲买 / ▼卖 = 区间中点操作信号（点击查看详情）</span><span class="anno" style="margin-left:14px">↔ 下方图表可左右滑动</span></div>
<div class="chart-scroll">
{svg_chart(name)}
</div>
<table>
<thead><tr><th>年份</th><th>年度最低</th><th>最低日</th><th>低点起</th><th>低点止</th><th>低点价格带</th><th>低点带浮动%</th><th>年度最高</th><th>最高日</th><th>高点起</th><th>高点止</th><th>高点价格带</th><th>高点带浮动%</th><th>低→高上涨%区间</th></tr></thead>
<tbody>{''.join(rows)}</tbody>
</table>
{caveat}
</div>"""

cards = "\n".join(html_table(n) for n in INDEX_ORDER)

# ---------------- 当前位置实时面板 ----------------
def live_panel_html():
    items = []
    for name in INDEX_ORDER:
        lv = RES["live"].get(name, {})
        if not lv:
            continue
        zone = lv["zone"]
        if zone == "低点带内":
            zc, zt = LOW_COLOR, "处于低点带（距年内最低收盘 ≤5%）"
        elif zone == "高点带内":
            zc, zt = HIGH_COLOR, "处于高点带（距年内最高收盘 ≥-5%）"
        else:
            zc, zt = "#FFD54F", "处于中间区（远离年内极值）"
        # 距带幅度可读性
        dl = lv["dist_low_pct"]; dh = lv["dist_high_pct"]
        items.append(f"""<div class="live-card">
  <div class="live-name">{name} <span class="code">({RES['indices'][name]['code']})</span></div>
  <div class="live-zone" style="color:{zc}">● {zone}</div>
  <div class="live-grid">
    <div><span class="k">今日收盘</span><span class="v">{lv['last_close']}</span><span class="d">{lv['last_date']}</span></div>
    <div><span class="k">年内最低收盘</span><span class="v green">{lv['ytd_low_close']}</span><span class="d">{lv['ytd_low_date']}</span></div>
    <div><span class="k">年内最高收盘</span><span class="v red">{lv['ytd_high_close']}</span><span class="d">{lv['ytd_high_date']}</span></div>
    <div><span class="k">距低点带</span><span class="v">{dl}%</span><span class="d">阈值 ≤{ALPHA*100:.0f}%</span></div>
    <div><span class="k">距高点带</span><span class="v">{dh}%</span><span class="d">阈值 ≥-{ALPHA*100:.0f}%</span></div>
    <div><span class="k">判断</span><span class="v" style="color:{zc}">{zt}</span></div>
  </div>
</div>""")
    disc = ('<p class="live-tip">实时读数说明：以「年初至今最低/最高收盘」为锚，沿用报告 ±5% 容忍带。'
            '今日收盘距年内最低收盘 ≤5% → 低点带内（可结合估值分位/情绪冰点分批建仓）；'
            '距年内最高收盘 ≥-5% → 高点带内（可分批止盈）。此为前瞻代理，非事后高低点。</p>')
    return f'<div class="live-panel"><h2>当前位置实时读数（截至 {RES["meta"]["gen_date"]}）</h2>{disc}<div class="live-row">{"".join(items)}</div></div>'

live_panel = live_panel_html()

def strategy_html():
    comp = STRAT["comprehensive"]
    etf_rows = ""
    for name in INDEX_ORDER:
        for j, e in enumerate(STRAT["etfs"][name]):
            etf_rows += (f"<tr><td class='l'>{name if j == 0 else ''}</td><td>{e['etf']}</td>"
                         f"<td class='l'>{e['etf_name']}</td>"
                         f"<td class='l note'>{e['note']}</td></tr>")
    etf_tbl = (f"<h3>一、代表 ETF 对照</h3>"
               f"<table class='etf-tbl'><thead><tr><th>指数</th><th>ETF代码</th><th>ETF名称</th><th>备注</th></tr></thead>"
               f"<tbody>{etf_rows}</tbody></table>")

    sum_items = (f"<div class='sum-box'>"
                 f"<div class='sum-item'><div class='k'>综合·平均收益率(极值·理论上限)</div><div class='v green'>{comp['avg_ext']}%</div></div>"
                 f"<div class='sum-item'><div class='k'>综合·平均收益率(区间中点·实操近似)</div><div class='v gold'>{comp['avg_mid']}%</div></div>"
                 f"<div class='sum-item'><div class='k'>综合·平均净收益(中点·扣费{FEE*100:.1f}%)</div><div class='v grey'>{comp['avg_mid_net']}%</div></div>"
                 f"</div>")

    cards_html = []
    for name in INDEX_ORDER:
        idx = STRAT["indices"][name]
        mini = (f"<div class='sum-box'>"
                f"<div class='sum-item'><div class='k'>{name} 平均(极值)</div><div class='v green'>{idx['avg_ext']}%</div></div>"
                f"<div class='sum-item'><div class='k'>{name} 平均(中点)</div><div class='v gold'>{idx['avg_mid']}%</div></div>"
                f"<div class='sum-item'><div class='k'>{name} 平均净(中点扣费)</div><div class='v grey'>{idx['avg_mid_net']}%</div></div>"
                f"<div class='sum-item'><div class='k'>可交易笔数</div><div class='v grey'>{idx['n_closed']}</div></div>"
                f"</div>")
        tr_rows = []
        for t in idx["trades"]:
            if t["open"]:
                tr_rows.append(
                    f"<tr><td>{t['buy_year']}</td><td>{t['buy_date']}</td><td>{t['buy_price_ext']}</td><td>{t['buy_band_mid']}</td>"
                    f"<td class='open'>持仓</td><td class='open'>—</td><td class='open'>—</td><td class='open'>—</td>"
                    f"<td class='open'>—</td><td class='open'>持仓未平仓</td><td class='open'>—</td></tr>")
            else:
                tr_rows.append(
                    f"<tr><td>{t['buy_year']}</td><td>{t['buy_date']}</td><td>{t['buy_price_ext']}</td><td>{t['buy_band_mid']}</td>"
                    f"<td>{t['sell_year']}</td><td>{t['sell_date']}</td><td>{t['sell_price_ext']}</td><td>{t['sell_band_mid']}</td>"
                    f"<td class='pos'>{t['r_ext']}%</td><td class='pos'>{t['r_mid']}%</td><td class='pos'>{round(t['r_mid']-FEE*100,2)}%</td></tr>")
        tbl = (f"<h3>{name} 逐年交易明细</h3>"
               f"<table class='trade-tbl'><thead><tr><th>买入年</th><th>买入日</th><th>买入点(极值)</th><th>买入点(中点)</th>"
               f"<th>卖出年</th><th>卖出日</th><th>卖出点(极值)</th><th>卖出点(中点)</th>"
               f"<th>收益率%(极值)</th><th>收益率%(中点)</th><th>净收益%(中点,扣费)</th></tr></thead>"
               f"<tbody>{''.join(tr_rows)}</tbody></table>")
        cards_html.append(f"<div class='card' style='margin-top:14px'>{mini}{tbl}</div>")
    per_index = "".join(cards_html)

    disc = ("""<h3>四、最具性价比的买卖方式（实操讨论）</h3>
<ul class="disc">
<li><b>后见之明不可复现：</b>本表高低点均为事后算出，真实无法精准抄底逃顶。上面「极值」是<b>理论天花板</b>，「区间中点」才是更接近可成交的近似——但中点仍是事后最优，实操需进一步打折。</li>
<li><b>分批进出（scale-in / scale-out）：</b>不要一次性买/卖。进到低点带后分 2–3 笔建仓，进到高点带后分批止盈，可平滑踏错时点与滑点。</li>
<li><b>用信号替代"知道高低点"：</b>真实场景靠规则触发——如价格跌破近年低点带下沿、或估值分位进入历史低位时开始买；价格进入近年高点带、估值分位偏高时分批卖。配合 20/60 日均线、成交额放大确认。</li>
<li><b>费用与摩擦：</b>ETF 管理费+双边滑点/佣金约 0.3%/轮已扣除；高频换仓会显著侵蚀收益，本策略一年最多 1–2 轮，费用可控。</li>
<li><b>不可预测变化（核心风险）：</b>2018 贸易战、2020 疫情、2024 政策牛等会让"跨年持有"在持有期遭遇大幅回撤；需设止损/回撤阈值（如买入后破位 -8% 减仓），并保留现金应对极端。</li>
<li><b>现金拖累与机会成本：</b>买入后到次年高点前资金被占用，期间现金几乎无收益；若判断当年"先高后低"（本样本约一半年份），应转为年内减仓/观望而非死扛跨年。</li>
<li><b>指数选择：</b>上证综指无主流大 ETF，本报告上证改用 <b>510050(上证50ETF) / 510300(沪深300ETF)</b> 作实操代理；二者跟踪的是上证50/沪深300，与综指存在风格偏离，回测点位口径仍为上证综指，二者走势吻合度见下方「指数与ETF重合图」。</li>
</ul>
<p class="note-small">说明：跨年持有规则下，2025 年低点为未平仓持仓（无 2026 数据），不计入均值。科创50 2019 仅基点记录，已剔除。</p>""")

    return (f"""<div class="strat">
<h2>低买高卖策略回测（实际操盘参考）</h2>
<div class="sub">规则：买入年低点区间买入 → 持有至次年高点区间卖出（跨年持有）；极值=理论上限，区间中点=实操近似；每轮往返扣费 {FEE*100:.1f}%。目的：验证"低买高卖"历史空间并剖析实操边界。</div>
{etf_tbl}
<h3>二、综合年盈利（三指数等权平均）</h3>
{sum_items}
<h3>三、分指数逐年交易明细</h3>
{per_index}
{disc}
</div>""")

strategy_section = strategy_html()

# prepare data for JS popup
popup_data = {}
for name in INDEX_ORDER:
    popup_data[name] = {}
    for y, v in RES["indices"][name]["years"].items():
        lp, hp, bp = band_pcts(v)
        popup_data[name][y] = {
            "annual_low": v["annual_low"], "annual_low_date": v["annual_low_date"],
            "annual_high": v["annual_high"], "annual_high_date": v["annual_high_date"],
            "low_window_start": v["low_window_start"], "low_window_end": v["low_window_end"],
            "high_window_start": v["high_window_start"], "high_window_end": v["high_window_end"],
            "low_band_low": v["low_band_low"], "low_band_high": v["low_band_high"],
            "high_band_low": v["high_band_low"], "high_band_high": v["high_band_high"],
            "low_pct": lp, "high_pct": hp, "band_to_band": bp,
        }

# ---------------- 指数 vs ETF 重合图（归一化到起点=100） ----------------

def overlay_chart(name):
    idx_series = RAW[name]["series"]
    idx_dates = [datetime.datetime.strptime(r["date"], "%Y-%m-%d").date() for r in idx_series]
    idx_by_date = {d: c for d, c in zip(idx_dates, (x["close"] for x in idx_series))}
    codes = ETF.get("index_map", {}).get(name, [])
    etf_lines = []
    for code in codes:
        s = ETF.get("etfs", {}).get(code, {}).get("series", [])
        if not s:
            continue
        bd = {datetime.datetime.strptime(r["date"], "%Y-%m-%d").date(): r["close"] for r in s}
        etf_lines.append((code, ETF["etfs"][code]["name"], bd))
    # 共同交易日（指数与所有 ETF 的交集）
    common = [d for d in idx_dates if d in idx_by_date and all(d in bd for _, _, bd in etf_lines)] or idx_dates
    start = common[0]
    base_idx = idx_by_date[start]
    base_etf = [bd[start] for _, _, bd in etf_lines]
    xs = [(d - start).days for d in common]
    xmax = max(xs) if xs else 1
    idx_norm = [idx_by_date[d] / base_idx * 100 for d in common]
    etf_norm = [(code, nm, [bd[d] / b0 * 100 for d in common]) for (code, nm, bd), b0 in zip(etf_lines, base_etf)]
    # ETF 数据缺失时（抓取被限流），退化为纯指数走势图
    etf_vals = [v for e in etf_norm for v in e[2]]
    pmin = min([min(idx_norm)] + etf_vals)
    pmax = max([max(idx_norm)] + etf_vals)
    pad = (pmax - pmin) * 0.08
    pLo, pHi = pmin - pad, pmax + pad
    W, H, left, right, top, bottom = 940, 360, 56, 16, 26, 58
    plotW, plotH = W - left - right, H - top - bottom
    def X(xd): return left + (xd / xmax * plotW) if xmax else left
    def Y(p): return top + plotH - (p - pLo) / (pHi - pLo) * plotH
    svg = [f'<svg viewBox="0 0 {W} {H}" width="100%" preserveAspectRatio="xMinYMin meet" font-family="-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif">']
    for k in range(5):
        p = pLo + (pHi - pLo) * k / 4
        y = Y(p)
        svg.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left+plotW}" y2="{y:.1f}" stroke="#1c3a5e" stroke-width="1"/>')
        svg.append(f'<text x="{left-8}" y="{y+3:.1f}" fill="#8fa8c4" font-size="10" text-anchor="end">{p:.0f}</text>')
    svg.append(f'<text x="{left-44}" y="{top-10}" fill="#8fa8c4" font-size="10" text-anchor="start">归一化(起点=100)</text>')
    cur = start.replace(month=1, day=1)
    if cur < start:
        cur = datetime.date(start.year + 1, 1, 1)
    while cur <= common[-1]:
        xd = (cur - start).days
        x = X(xd)
        svg.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top+plotH}" stroke="#15324f" stroke-width="1"/>')
        svg.append(f'<text x="{x:.1f}" y="{top+plotH+18}" fill="#8fa8c4" font-size="10" text-anchor="middle">{cur.year}</text>')
        cur = datetime.date(cur.year + 1, 1, 1)
    pts = " ".join(f"{X(xd):.1f},{Y(p):.1f}" for xd, p in zip(xs, idx_norm))
    svg.append(f'<polyline points="{pts}" fill="none" stroke="#5fa8ff" stroke-width="2" stroke-linejoin="round"/>')
    for i, (code, nm, vals) in enumerate(etf_norm):
        col = ETF_LINE_COLORS[i % len(ETF_LINE_COLORS)]
        pts = " ".join(f"{X(xd):.1f},{Y(p):.1f}" for xd, p in zip(xs, vals))
        svg.append(f'<polyline points="{pts}" fill="none" stroke="{col}" stroke-width="1.6" stroke-dasharray="5 3" stroke-linejoin="round" opacity="0.95"/>')
    lx = left + 6
    svg.append(f'<text x="{lx}" y="{top+12}" fill="#5fa8ff" font-size="11" font-weight="600">● 指数</text>')
    ox = lx + 52
    for i, (code, nm, _) in enumerate(etf_norm):
        col = ETF_LINE_COLORS[i % len(ETF_LINE_COLORS)]
        svg.append(f'<text x="{ox}" y="{top+12}" fill="{col}" font-size="11">┄ {nm}({code})</text>')
        ox += 160
    svg.append(f'<text x="{left+plotW}" y="{top+plotH+42}" fill="#6f8bab" font-size="10" text-anchor="end">起点 {start} 归一化=100 ｜ 数值越高=相对起点涨幅越大</text>')
    svg.append('</svg>')
    return "".join(svg)

OVERLAY_CAVEAT = {
    "上证指数": "上证综指无对应大 ETF，叠加的 510050(上证50)/510300(沪深300) 跟踪的是不同指数，二者与综指分叉属正常风格偏离，非跟踪误差。",
    "创业板指": "159915 直接跟踪创业板指，理论应基本重合；局部微小偏离来自 ETF 申赎摩擦与现金拖累。",
    "科创50":   "588000 直接跟踪科创50，理论应基本重合；ETF 首交易日为 2020-11-16，故重合图自该日起。" ,
}
overlay_section = (
    "<div class='overlay-sec'>"
    "<h2>指数与 ETF 走势重合图（归一化对比）</h2>"
    "<div class='sub'>同一时间轴、各自以起点=100 归一化，直接看 ETF 是否跟住指数；上证叠加 510050 / 510300 两个实操代理。重合度越高，代理越可靠。</div>"
    + "".join(
        f"<div class='card'><h2>{name} <span class='code'>vs ETF</span></h2>{overlay_chart(name)}"
        f"<p class='caveat'>{OVERLAY_CAVEAT.get(name, '')}</p></div>"
        for name in INDEX_ORDER)
    + "</div>")

html = f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>A股三大指数 逐年低点/高点区间报表</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#0a1929;color:#e6f0fa;font-family:-apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;line-height:1.5}}
.wrap{{max-width:1000px;margin:0 auto;padding:28px 18px 60px}}
header h1{{font-size:22px;margin:0 0 6px;color:#fff}}
header .sub{{color:#8fa8c4;font-size:13px;margin-bottom:4px}}
header .meta{{color:#6f8bab;font-size:12px}}
.cards{{margin-top:18px}}
.card{{background:#102a43;border:1px solid #1c3a5e;border-radius:12px;padding:18px 18px 22px;margin-bottom:22px;box-shadow:0 4px 18px rgba(0,0,0,.25)}}
.card h2{{margin:0 0 10px;font-size:18px;color:#fff}}
.card .code{{color:#5fa8ff;font-size:13px;font-weight:400}}
.legend{{font-size:12px;color:#9fb6d2;margin-bottom:8px}}
.legend .dot{{display:inline-block;width:10px;height:10px;border-radius:3px;vertical-align:middle;margin-right:3px}}
.dot.low{{background:{LOW_COLOR}}} .dot.high{{background:{HIGH_COLOR}}}
.legend .anno{{margin-left:8px;color:#7d97b6}}
.chart-scroll{{width:100%;overflow-x:auto;overflow-y:hidden;margin:6px 0 14px;background:#0c2238;border-radius:8px;border:1px solid #1c3a5e;box-shadow:0 4px 18px rgba(0,0,0,.25)}}
.chart-scroll svg{{display:block;background:#0c2238;border-radius:0;border:0;margin:0;padding:0}}
table{{width:100%;border-collapse:collapse;font-size:12.5px}}
th,td{{padding:7px 6px;border-bottom:1px solid #1c3a5e;text-align:center}}
th{{background:#16395c;color:#cfe0f2;position:sticky;top:0;font-weight:600}}
td.num{{color:{LOW_COLOR};font-weight:600}} td.num.hi{{color:{HIGH_COLOR}}}
td.band{{color:#cfe0f2}}
td.pct{{color:#a8c5e2}}
td.key{{background:rgba(255,193,7,0.12);color:#ffd54f;font-weight:600}}
tbody tr:hover{{background:#14304d}}
tbody tr.hl{{background:#1f4e79 !important}}
tbody tr.hl td.key{{background:rgba(255,193,7,0.22)}}
.caveat{{color:#ffcf6b;font-size:12px;margin-top:10px}}
footer{{margin-top:30px;color:#6f8bab;font-size:12px;border-top:1px solid #1c3a5e;padding-top:14px}}
footer b{{color:#ff8a8a}}
#popup-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:100;justify-content:center;align-items:center}}
#popup-overlay.active{{display:flex}}
#point-popup{{background:#102a43;border:1px solid #1c3a5e;border-radius:12px;padding:0;width:380px;max-width:92vw;box-shadow:0 10px 34px rgba(0,0,0,.5);color:#e6f0fa}}
.popup-header{{display:flex;justify-content:space-between;align-items:center;padding:14px 18px;border-bottom:1px solid #1c3a5e}}
.popup-header h3{{margin:0;font-size:15px;color:#fff}}
.close-btn{{background:none;border:none;color:#8fa8c4;font-size:22px;line-height:1;cursor:pointer}}
.popup-body{{padding:10px 18px 16px}}
.popup-row{{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1c3a5e;font-size:13px}}
.popup-row:last-child{{border-bottom:none}}
.popup-row .label{{color:#8fa8c4}}
.popup-row .value{{color:#fff;font-weight:600}}
.popup-row .value.green{{color:{LOW_COLOR}}}
.popup-row .value.red{{color:{HIGH_COLOR}}}
.strat{{margin-top:26px}}
.strat h2{{font-size:19px;color:#fff;margin:0 0 4px}}
.strat .sub{{color:#8fa8c4;font-size:12.5px;margin-bottom:14px}}
.strat h3{{font-size:15px;color:#cfe0f2;margin:18px 0 8px}}
.etf-tbl,.trade-tbl{{width:100%;border-collapse:collapse;font-size:12px}}
.etf-tbl th,.etf-tbl td,.trade-tbl th,.trade-tbl td{{padding:6px 7px;border-bottom:1px solid #1c3a5e;text-align:center}}
.etf-tbl th,.trade-tbl th{{background:#16395c;color:#cfe0f2;font-weight:600}}
.etf-tbl td.l,.trade-tbl td.l{{text-align:left}}
.etf-tbl td.note{{color:#8fa8c4;font-size:11.5px}}
.trade-tbl td.open{{color:#7d97b6}}
.trade-tbl td.pos{{color:{LOW_COLOR};font-weight:600}}
.sum-box{{display:flex;flex-wrap:wrap;gap:12px;margin:8px 0 6px}}
.sum-item{{flex:1;min-width:170px;background:#0e2438;border:1px solid #1c3a5e;border-radius:10px;padding:12px 14px}}
.sum-item .k{{color:#8fa8c4;font-size:12px}}
.sum-item .v{{color:#fff;font-size:21px;font-weight:700;margin-top:3px}}
.sum-item .v.green{{color:{LOW_COLOR}}} .sum-item .v.gold{{color:#ffd54f}} .sum-item .v.grey{{color:#9fb6d2}}
.disc{{font-size:13px;color:#cfe0f2;line-height:1.75;padding-left:18px;margin:6px 0 0}}
.disc li{{margin-bottom:8px}}
.disc b{{color:#fff}}
.note-small{{color:#7d97b6;font-size:11.5px;margin-top:8px}}
.overlay-sec{{margin-top:26px}}
.overlay-sec h2{{font-size:19px;color:#fff;margin:0 0 4px}}
.overlay-sec .sub{{color:#8fa8c4;font-size:12.5px;margin-bottom:14px}}
.overlay-sec .card{{background:#102a43;border:1px solid #1c3a5e;border-radius:12px;padding:18px 18px 22px;margin-bottom:22px;box-shadow:0 4px 18px rgba(0,0,0,.25)}}
.overlay-sec .card h2{{margin:0 0 8px;font-size:17px;color:#fff}}
.overlay-sec svg{{display:block;background:#0c2238;border-radius:8px;border:1px solid #1c3a5e;margin:4px 0 6px}}
.live-panel{{background:#0e2438;border:1px solid #1c3a5e;border-radius:12px;padding:16px 18px 18px;margin:16px 0 8px;box-shadow:0 4px 18px rgba(0,0,0,.25)}}
.live-panel h2{{margin:0 0 8px;font-size:17px;color:#fff}}
.live-tip{{color:#8fa8c4;font-size:12px;margin:0 0 12px;line-height:1.6}}
.live-row{{display:flex;flex-wrap:wrap;gap:14px}}
.live-card{{flex:1;min-width:260px;background:#102a43;border:1px solid #1c3a5e;border-radius:10px;padding:12px 14px}}
.live-name{{font-size:14px;color:#fff;font-weight:600}}
.live-name .code{{color:#5fa8ff;font-size:12px;font-weight:400}}
.live-zone{{font-size:14px;font-weight:700;margin:4px 0 8px}}
.live-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px 14px}}
.live-grid > div{{display:flex;flex-direction:column}}
.live-grid .k{{color:#7d97b6;font-size:11px}}
.live-grid .v{{color:#e6f0fa;font-size:15px;font-weight:700}}
.live-grid .v.green{{color:{LOW_COLOR}}} .live-grid .v.red{{color:{HIGH_COLOR}}}
.live-grid .d{{font-size:10px;color:#6f8bab}}
</style></head>
<body><div class="wrap">
<header>
<h1>A股三大指数 · 逐年低点/高点区间报表</h1>
<div class="sub">上证指数 · 创业板指 · 科创50 ｜ 2016 至 {END_DISP}（动态更新）</div>
<div class="meta">生成日期 {GEN_DATE} ｜ 数据源：东方财富日K线(前复权) ｜ 区间定义：±{ALPHA:.0%} 容忍带主区间法</div>
</header>
{live_panel}
<div class="cards">{cards}</div>
{overlay_section}
{strategy_section}
<div id="popup-overlay" onclick="if(event.target===this)hidePopup()">
  <div id="point-popup">
    <div class="popup-header"><h3 id="popup-title"></h3><button class="close-btn" onclick="hidePopup()">×</button></div>
    <div class="popup-body">
      <div class="popup-row"><span class="label">日期</span><span class="value" id="popup-date"></span></div>
      <div class="popup-row"><span class="label">点位</span><span class="value" id="popup-price"></span></div>
      <div class="popup-row"><span class="label">区间起止</span><span class="value" id="popup-window"></span></div>
      <div class="popup-row"><span class="label">价格带</span><span class="value" id="popup-band"></span></div>
      <div class="popup-row"><span class="label">区间浮动%</span><span class="value" id="popup-band-pct"></span></div>
      <div class="popup-row"><span class="label">低→高上涨%区间</span><span class="value" id="popup-b2b"></span></div>
    </div>
  </div>
</div>
<footer>
<p><b>方法说明：</b>年度最低/最高 = 全年每日最低/最高价极值；低点区间 = 以年最低点交易日为锚向两侧扩展至收盘价突破 年最低×1.05 的连续时段，取起止日期与期间价格带；高点区间对称(阈值 年最高×0.95)。仅取主区间。</p>
<p><b>价格带浮动%：</b>低点带浮动%=(低点带上沿−下沿)/下沿；高点带浮动%=(高点带上沿−下沿)/下沿；低→高上涨%区间=(高点带中点−低点带中点)/低点带中点。</p>
<p><b>免责声明：</b>本报表仅基于公开行情数据的量化整理，不构成任何投资建议。市场有风险，决策需谨慎。</p>
</footer>
</div>
<script>
const INDEX_DATA = {json.dumps(popup_data, ensure_ascii=False)};
function highlightRow(name, year){{
  var card = document.getElementById('card-'+name);
  if (!card) return;
  card.querySelectorAll('tbody tr').forEach(function(r){{ r.classList.remove('hl'); }});
  var row = document.getElementById('row-'+name+'-'+year);
  if (row){{
    row.classList.add('hl');
    row.scrollIntoView({{behavior:'smooth', block:'nearest'}});
  }}
}}
function clearRow(name){{
  var card = document.getElementById('card-'+name);
  if (card) card.querySelectorAll('tbody tr.hl').forEach(function(r){{ r.classList.remove('hl'); }});
}}
function showPopup(name, year, type){{
  var d = INDEX_DATA[name] && INDEX_DATA[name][year];
  if (!d) return;
  var title = name + ' ' + year + ' 年' + (type === 'low' ? '最低点' : '最高点');
  var date = type === 'low' ? d.annual_low_date : d.annual_high_date;
  var price = type === 'low' ? d.annual_low : d.annual_high;
  var win = type === 'low' ? (d.low_window_start + ' ~ ' + d.low_window_end) : (d.high_window_start + ' ~ ' + d.high_window_end);
  var band = type === 'low' ? (d.low_band_low.toFixed(2) + ' ~ ' + d.low_band_high.toFixed(2)) : (d.high_band_low.toFixed(2) + ' ~ ' + d.high_band_high.toFixed(2));
  var bandPct = type === 'low' ? d.low_pct : d.high_pct;
  document.getElementById('popup-title').textContent = title;
  document.getElementById('popup-date').textContent = date;
  document.getElementById('popup-price').textContent = price.toFixed(2);
  document.getElementById('popup-price').className = 'value ' + (type === 'low' ? 'green' : 'red');
  document.getElementById('popup-window').textContent = win;
  document.getElementById('popup-band').textContent = band;
  document.getElementById('popup-band-pct').textContent = bandPct + '%';
  document.getElementById('popup-b2b').textContent = d.band_to_band + '%';
  document.getElementById('popup-overlay').classList.add('active');
}}
function hidePopup(){{
  document.getElementById('popup-overlay').classList.remove('active');
}}
document.addEventListener('keydown', function(e){{ if (e.key === 'Escape') hidePopup(); }});
</script>
</body></html>"""

html_path = os.path.join(HERE, "index_low_high_report.html")
with open(html_path, "w", encoding="utf-8") as f:
    f.write(html)
print("saved html ->", html_path)
